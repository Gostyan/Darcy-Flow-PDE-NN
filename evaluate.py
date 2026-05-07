"""
evaluate.py — Comprehensive evaluation of U-Net, FNO, and PINN-U-Net on 2D Darcy Flow.

Metrics reported (mainstream PDE solution benchmarks)
------------------------------------------------------
  RMSE             Root Mean Squared Error
  MAE              Mean Absolute Error
  Rel-L2           Relative L2 / nRMSE  (||u_pred - u||_2 / ||u||_2)
  Max-Err          L-infinity norm  (max |u_pred - u|)
  PDE-Res RMSE     Root-mean-square PDE residual  |-div(a grad u) - 1|
  BC-RMSE          Boundary condition RMSE  ||u_pred||_{dOmega}

Outputs
-------
  results/
    metrics_summary.xlsx                       — full metrics table (Excel)
    Figure_1_Darcy_Field_Comparison_Sample_{i}.pdf  — 2-D field + error maps
    Figure_2_Training_Loss_and_Validation_Curves.pdf  — training curves
    Figure_3_Model_Performance_Metrics_Comparison.pdf — bar chart

Usage
-----
    python evaluate.py
    python evaluate.py --num_vis 5 --results_dir results
"""

from __future__ import annotations

import argparse
import io
import json
import tarfile
from pathlib import Path
from typing import Optional

import matplotlib
matplotlib.use("Agg")                           # headless rendering
import matplotlib.pyplot as plt
import matplotlib.gridspec as gridspec
from matplotlib import font_manager as _fm
import numpy as np
import torch
from torch.utils.data import DataLoader
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from models import FNO2d, UNet2d
from utils import DarcyDataset, compute_metrics

# ---------------------------------------------------------------------------
# Publication-quality style (STIX ≈ Times New Roman for math/text)
# ---------------------------------------------------------------------------

_has_times = any("Times New Roman" in f.name for f in _fm.fontManager.ttflist)
_SERIF     = ["Times New Roman"] if _has_times else ["STIXGeneral", "DejaVu Serif"]
_FONT_NAME = _SERIF[0]

plt.rcParams.update({
    "font.family":       "serif",
    "font.serif":        _SERIF,
    "mathtext.fontset":  "stix",
    "font.size":         9,
    "axes.titlesize":    9,
    "axes.labelsize":    9,
    "xtick.labelsize":   8,
    "ytick.labelsize":   8,
    "legend.fontsize":   8,
    "savefig.dpi":       300,
    "axes.linewidth":    0.8,
    "xtick.major.width": 0.8,
    "ytick.major.width": 0.8,
    "xtick.direction":   "in",
    "ytick.direction":   "in",
    "figure.facecolor":  "white",
    "axes.facecolor":    "white",
})

# Metric columns shown in the Excel / terminal table
METRIC_KEYS = ["rmse", "mae", "rel_l2", "max_err", "pde_residual_rmse", "boundary_err"]
METRIC_LABELS = {
    "rmse":              "RMSE",
    "mae":               "MAE",
    "rel_l2":            "Rel-L2",
    "max_err":           "Max-Err (L∞)",
    "pde_residual_rmse": "PDE-Res RMSE",
    "boundary_err":      "BC-RMSE",
}

# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def get_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="Evaluate Darcy Flow models")
    p.add_argument("--data_path",      type=str, default="dataset/2D_DarcyFlow_beta1.0_Train.hdf5")
    p.add_argument("--batch_size",     type=int, default=8)
    p.add_argument("--reduced_res",    type=int, default=1)
    p.add_argument("--num_samples",    type=int, default=-1,
                   help="Max test samples (-1 = all)")
    p.add_argument("--results_dir",    type=str, default="results")
    p.add_argument("--num_vis",        type=int, default=3,
                   help="Number of test samples to visualise (default 3)")
    p.add_argument("--vis_seed",       type=int, default=None,
                   help="Random seed for selecting visualisation samples "
                        "(default: None = use first --num_vis samples)")
    # checkpoint paths
    p.add_argument("--ckpt_unet",      type=str, default="checkpoints/unet/best_model.pt")
    p.add_argument("--ckpt_fno",       type=str, default="checkpoints/fno/best_model.pt")
    p.add_argument("--ckpt_pinn_unet", type=str, default="checkpoints/pinn_unet/best_model.pt")
    # history paths (for training-curve plots)
    p.add_argument("--hist_unet",      type=str, default="checkpoints/unet/history.json")
    p.add_argument("--hist_fno",       type=str, default="checkpoints/fno/history.json")
    p.add_argument("--hist_pinn_unet", type=str, default="checkpoints/pinn_unet/history.json")
    # model hyper-parameters (must match training)
    p.add_argument("--init_features",  type=int, default=32)
    p.add_argument("--fno_modes",      type=int, default=12)
    p.add_argument("--fno_width",      type=int, default=32)
    return p.parse_args()


# ---------------------------------------------------------------------------
# Model loaders
# ---------------------------------------------------------------------------

def _detect_fno_width(ckpt_path: str) -> Optional[int]:
    """Read fc0.weight shape from a .pt checkpoint to auto-detect FNO width."""
    try:
        sd = torch.load(ckpt_path, map_location="cpu", weights_only=True)
        sd = sd.get("model", sd)
        return int(sd["fc0.weight"].shape[0])
    except Exception:
        return None


def load_unet(ckpt_path: str, init_features: int, device) -> Optional[UNet2d]:
    if not Path(ckpt_path).exists():
        print(f"  [N/A] UNet checkpoint not found: {ckpt_path}")
        return None
    model = UNet2d(in_channels=1, out_channels=1, init_features=init_features).to(device)
    sd = torch.load(ckpt_path, map_location=device, weights_only=True)
    sd = sd.get("model", sd)
    model.load_state_dict(sd)
    model.eval()
    return model


def load_fno(ckpt_path: str, modes: int, width: int, device) -> Optional[FNO2d]:
    if not Path(ckpt_path).exists():
        print(f"  [N/A] FNO checkpoint not found: {ckpt_path}")
        return None
    detected = _detect_fno_width(ckpt_path)
    if detected and detected != width:
        print(f"  [INFO] FNO width auto-detected: {detected} (override --fno_width {width})")
        width = detected
    model = FNO2d(num_channels=1, modes1=modes, modes2=modes,
                  width=width, initial_step=1).to(device)
    sd = torch.load(ckpt_path, map_location=device, weights_only=True)
    sd = sd.get("model", sd)
    model.load_state_dict(sd)
    model.eval()
    return model


# ---------------------------------------------------------------------------
# Evaluation
# ---------------------------------------------------------------------------

@torch.no_grad()
def evaluate_model(model, loader, device, model_type: str) -> dict[str, float]:
    """Accumulate metrics over the full test set."""
    accum  = {k: 0.0 for k in METRIC_KEYS + ["mse", "pde_residual"]}
    n_total = 0

    for a, u, grid in loader:
        a    = a.to(device)
        u    = u.to(device)
        grid = grid.to(device)

        if model_type == "unet":
            u_pred = model(a.unsqueeze(1)).squeeze(1)
        elif model_type == "fno":
            grid0  = grid[0] if grid.dim() == 4 else grid
            u_pred = model(a.unsqueeze(-1), grid0).squeeze(-1).squeeze(-1)
        else:
            raise ValueError(f"Unknown model_type: {model_type}")

        m = compute_metrics(a, u_pred.float(), u)
        B = a.size(0)
        for k in accum:
            if k in m:
                accum[k] += m[k] * B
        n_total += B

    return {k: v / n_total for k, v in accum.items()}


# ---------------------------------------------------------------------------
# Excel export
# ---------------------------------------------------------------------------

def export_excel(results: dict[str, Optional[dict]], save_path: Path) -> None:
    """Write metrics to a styled Excel workbook."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Metrics"

    # styles
    hdr_font  = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    hdr_fill  = PatternFill("solid", fgColor="2F5496")
    model_fill = PatternFill("solid", fgColor="D9E1F2")
    na_font   = Font(name="Calibri", italic=True, color="999999")
    center    = Alignment(horizontal="center", vertical="center")
    thin      = Side(style="thin", color="BFBFBF")
    border    = Border(left=thin, right=thin, top=thin, bottom=thin)

    # header row
    headers = ["Model"] + [METRIC_LABELS[k] for k in METRIC_KEYS]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = center
        cell.border    = border

    # data rows
    for row_idx, (name, metrics) in enumerate(results.items(), 2):
        # model name cell
        c = ws.cell(row=row_idx, column=1, value=name)
        c.font      = Font(name="Calibri", bold=True, size=10)
        c.fill      = model_fill
        c.alignment = center
        c.border    = border

        for col_idx, k in enumerate(METRIC_KEYS, 2):
            if metrics is None or k not in metrics:
                cell = ws.cell(row=row_idx, column=col_idx, value="N/A")
                cell.font = na_font
            else:
                val  = metrics[k]
                cell = ws.cell(row=row_idx, column=col_idx, value=round(float(val), 6))
                cell.number_format = "0.000000"
            cell.alignment = center
            cell.border    = border

    # column widths
    ws.column_dimensions["A"].width = 14
    for col in range(2, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 16

    # freeze header
    ws.freeze_panes = "A2"

    wb.save(save_path)
    print(f"Excel saved: {save_path}")


# ---------------------------------------------------------------------------
# Figure 1 — 2D Field Comparison
# ---------------------------------------------------------------------------

@torch.no_grad()
def save_field_comparison(
    models_info: dict,
    sample_a: torch.Tensor,
    sample_u: torch.Tensor,
    sample_grid: torch.Tensor,
    device,
    sample_idx: int,
    save_path: Path,
) -> None:
    """
    2-row figure:
      Row 0 — Permeability | Reference | Model predictions
      Row 1 — (empty) | (empty) | Absolute error fields
    """
    names  = list(models_info.keys())
    n_mdl  = len(names)
    ncols  = 2 + n_mdl
    nrows  = 2

    fig_w = 2.6 * ncols
    fig_h = 2.6 * nrows + 1.0        # extra height for caption
    fig, axes = plt.subplots(nrows, ncols,
                             figsize=(fig_w, fig_h),
                             constrained_layout=False)
    fig.subplots_adjust(left=0.04, right=0.98, top=0.90,
                        bottom=0.12, wspace=0.35, hspace=0.45)

    a_np = sample_a.numpy()
    u_np = sample_u.numpy()

    # collect predictions
    preds, errors = {}, {}
    for name, (model, mtype) in models_info.items():
        if model is None:
            preds[name] = errors[name] = None
            continue
        a_  = sample_a.to(device).unsqueeze(0)
        g_  = sample_grid.to(device)
        if mtype == "unet":
            u_p = model(a_.unsqueeze(1)).squeeze().cpu().numpy()
        else:
            u_p = model(a_.unsqueeze(-1), g_).squeeze().cpu().numpy()
        preds[name]  = u_p
        errors[name] = np.abs(u_p - u_np)

    u_vmin, u_vmax = float(u_np.min()), float(u_np.max())
    e_vmax = max(
        (e.max() for e in errors.values() if e is not None),
        default=1.0
    )

    def _fmt_ax(ax, xlabel=True, ylabel=True):
        if xlabel: ax.set_xlabel("$x$", labelpad=2)
        if ylabel: ax.set_ylabel("$y$", labelpad=2)
        ax.tick_params(which="both", labelbottom=xlabel, labelleft=ylabel)

    def _cbar(im, ax):
        cb = plt.colorbar(im, ax=ax, fraction=0.046, pad=0.04)
        cb.ax.tick_params(labelsize=7)

    # ── Row 0: input + reference + predictions ───────────────────────────
    im = axes[0, 0].imshow(a_np, origin="lower", cmap="viridis")
    axes[0, 0].set_title("Permeability $a(x,y)$")
    _fmt_ax(axes[0, 0])
    _cbar(im, axes[0, 0])

    im = axes[0, 1].imshow(u_np, origin="lower", cmap="RdBu_r",
                            vmin=u_vmin, vmax=u_vmax)
    axes[0, 1].set_title("Reference $u(x,y)$")
    _fmt_ax(axes[0, 1])
    _cbar(im, axes[0, 1])

    for j, name in enumerate(names):
        ax = axes[0, 2 + j]
        if preds[name] is None:
            ax.text(0.5, 0.5, "N/A", ha="center", va="center",
                    transform=ax.transAxes, fontsize=11)
            ax.set_title(name)
            ax.axis("off")
        else:
            rl2 = np.linalg.norm(preds[name] - u_np) / (np.linalg.norm(u_np) + 1e-8)
            im = ax.imshow(preds[name], origin="lower", cmap="RdBu_r",
                           vmin=u_vmin, vmax=u_vmax)
            ax.set_title(f"{name}\nRel-$L_2$={rl2:.3e}")
            _fmt_ax(ax)
            _cbar(im, ax)

    # ── Row 1: error fields (first two cols hidden) ───────────────────────
    for j in range(2):
        axes[1, j].axis("off")

    for j, name in enumerate(names):
        ax = axes[1, 2 + j]
        if errors[name] is None:
            ax.text(0.5, 0.5, "N/A", ha="center", va="center",
                    transform=ax.transAxes, fontsize=11)
            ax.set_title(f"{name} — Abs. Error")
            ax.axis("off")
        else:
            mae_val = errors[name].mean()
            im = ax.imshow(errors[name], origin="lower", cmap="hot_r",
                           vmin=0, vmax=e_vmax)
            ax.set_title(f"{name} — |error|\nMAE={mae_val:.3e}")
            _fmt_ax(ax)
            _cbar(im, ax)

    # ── Figure-level caption ──────────────────────────────────────────────
    fig.suptitle(
        f"Darcy Flow Solution — Test Sample #{sample_idx + 1}",
        fontsize=10, fontweight="bold", y=0.97,
    )
    last_col = 2 + n_mdl
    caption = (
        f"Figure 1-{{sample_idx+1}}. Comparison of predicted pressure fields "
        f"and absolute point-wise errors for a 2-D steady-state Darcy flow "
        f"test sample (index {{sample_idx+1}}). "
        "Column 1: permeability field a(x,y). "
        "Column 2: reference solution u(x,y). "
        f"Columns 3-{{last_col}}: model predictions (top row) and absolute errors (bottom row). "
        "All solution panels share the same diverging color scale; "
        "error panels share the same sequential scale."
    )

    fig.text(0.5, 0.01, caption, ha="center", va="bottom",
             fontsize=7, style="italic",
             wrap=True, transform=fig.transFigure)

    plt.savefig(save_path, dpi=300, bbox_inches="tight")
    plt.savefig(save_path.with_suffix(".png"), dpi=150, bbox_inches="tight")
    plt.close(fig)
    print(f"  Saved: {save_path.name}  +  {save_path.stem}.png")


# ---------------------------------------------------------------------------
# Figure 2 — Training Curves
# ---------------------------------------------------------------------------

def save_training_curves(
    hist_paths: dict[str, str],
    save_path: Path,
) -> None:
    """Plot train MSE loss and validation Rel-L2 for all models."""
    histories = {}
    for name, hp in hist_paths.items():
        p = Path(hp)
        if p.exists():
            try:
                histories[name] = json.loads(p.read_text())
            except Exception:
                pass

    fig, axes = plt.subplots(1, 2, figsize=(9, 3.8))
    fig.subplots_adjust(left=0.09, right=0.97, top=0.88,
                        bottom=0.18, wspace=0.32)

    colors  = ["#5588bb", "#cc7766", "#66aa88"]
    markers = ["o", "s", "^"]

    if not histories:
        for ax in axes:
            ax.text(0.5, 0.5,
                    "No training history found.\nRun training scripts first.",
                    ha="center", va="center",
                    transform=ax.transAxes, fontsize=10, color="gray")
    else:
        for i, (name, hist) in enumerate(histories.items()):
            epochs     = hist.get("epoch", [])
            train_mse  = hist.get("train_mse", [])
            val_rel_l2 = hist.get("val_rel_l2", [])
            kw = dict(color=colors[i % len(colors)],
                      marker=markers[i % len(markers)],
                      markersize=3, linewidth=1.2, label=name)
            if train_mse:
                axes[0].semilogy(epochs, train_mse, **kw)
            if val_rel_l2:
                axes[1].semilogy(epochs, val_rel_l2, **kw)

    axes[0].set_xlabel("Epoch")
    axes[0].set_ylabel("Training MSE Loss")
    axes[0].set_title("(a) Training Loss")
    if histories: axes[0].legend(framealpha=0.8)
    axes[0].tick_params(which="both", direction="in")
    axes[0].grid(True, which="major", linestyle="--", linewidth=0.5, alpha=0.6)

    axes[1].set_xlabel("Epoch")
    axes[1].set_ylabel("Relative $L_2$ Error")
    axes[1].set_title("(b) Validation Rel-$L_2$")
    if histories: axes[1].legend(framealpha=0.8)
    axes[1].tick_params(which="both", direction="in")
    axes[1].grid(True, which="major", linestyle="--", linewidth=0.5, alpha=0.6)

    fig.suptitle("Training Progress on 2-D Darcy Flow",
                 fontsize=10, fontweight="bold")
    caption = (
        "Figure 2. Training loss (MSE) and validation relative $L_2$ error curves "
        "for U-Net, FNO, and PINN-U-Net on the 2-D steady-state Darcy flow dataset. "
        "Both axes use a logarithmic scale. "
        "Curves are plotted at every evaluation interval."
    )
    fig.text(0.5, 0.01, caption, ha="center", va="bottom",
             fontsize=7, style="italic", transform=fig.transFigure)

    plt.savefig(save_path, dpi=300, bbox_inches="tight")
    plt.savefig(save_path.with_suffix(".png"), dpi=150, bbox_inches="tight")
    plt.close(fig)
    print(f"  Saved: {save_path.name}  +  {save_path.stem}.png")


# ---------------------------------------------------------------------------
# Figure 3 — Metrics Bar Chart
# ---------------------------------------------------------------------------

def save_metrics_bar(results: dict[str, Optional[dict]], save_path: Path) -> None:
    """Grouped bar chart comparing models across all metrics."""
    valid = {k: v for k, v in results.items() if v is not None}
    if not valid:
        print("  [skip] No valid model results for bar chart.")
        return

    metrics_to_plot = ["rmse", "mae", "rel_l2", "max_err", "pde_residual_rmse", "boundary_err"]
    xlabels = [METRIC_LABELS[k] for k in metrics_to_plot]
    model_names = list(valid.keys())
    n_models = len(model_names)
    n_metrics = len(metrics_to_plot)

    colors  = ["#5588bb", "#cc7766", "#66aa88"]
    x       = np.arange(n_metrics)
    width   = 0.22
    offsets = np.linspace(-(n_models - 1) / 2, (n_models - 1) / 2, n_models) * width

    fig, ax = plt.subplots(figsize=(10, 4.2))
    fig.subplots_adjust(left=0.08, right=0.97, top=0.88, bottom=0.26)

    for i, (name, met) in enumerate(valid.items()):
        vals = [met.get(k, 0.0) for k in metrics_to_plot]
        bars = ax.bar(x + offsets[i], vals, width,
                      label=name, color=colors[i % len(colors)],
                      alpha=0.85, edgecolor="white", linewidth=0.5)
        # value labels on bars
        for bar, v in zip(bars, vals):
            ax.text(bar.get_x() + bar.get_width() / 2,
                    bar.get_height() * 1.10,
                    f"{v:.2e}",
                    ha="center", va="bottom", fontsize=6, rotation=90,
                    clip_on=False)

    # N/A annotation for missing models
    for name in results:
        if results[name] is None:
            ax.bar([0], [0], label=f"{name} (N/A)",
                   color="lightgray", hatch="//", edgecolor="gray")

    ax.set_yscale("log")
    ax.set_xticks(x)
    ax.set_xticklabels(xlabels, rotation=30, ha="right", fontsize=8)
    ax.set_ylabel("Metric Value (log scale)")
    ax.set_title("Model Performance Comparison on 2-D Darcy Flow Test Set",
                 fontsize=10, fontweight="bold")
    ax.legend(framealpha=0.85, fontsize=8)
    ax.tick_params(which="both", direction="in")
    ax.grid(True, axis="y", which="major", linestyle="--", linewidth=0.5, alpha=0.6)

    # expand y-axis headroom to prevent rotated labels from overflowing
    y_lo, y_hi = ax.get_ylim()
    ax.set_ylim(y_lo, y_hi * 8)

    caption = (
        "Figure 3. Comparison of U-Net, FNO, and PINN-U-Net on six evaluation metrics "
        "for the 2-D steady-state Darcy flow test set ($N=1000$). "
        "RMSE: root mean squared error; MAE: mean absolute error; "
        "Rel-$L_2$: relative $L_2$ norm; Max-Err: $L_\\infty$ norm; "
        "PDE-Res: RMSE of the finite-difference PDE residual; "
        "BC-RMSE: boundary condition satisfaction error. "
        "Lower values indicate better performance."
    )
    fig.text(0.5, 0.01, caption, ha="center", va="bottom",
             fontsize=7, style="italic", transform=fig.transFigure)

    plt.savefig(save_path, dpi=300, bbox_inches="tight")
    plt.savefig(save_path.with_suffix(".png"), dpi=150, bbox_inches="tight")
    plt.close(fig)
    print(f"  Saved: {save_path.name}  +  {save_path.stem}.png")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main(args: argparse.Namespace) -> None:
    device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
    print(f"\nDevice: {device}")
    print(f"Font: {_FONT_NAME}\n")

    results_dir = Path(args.results_dir)
    results_dir.mkdir(parents=True, exist_ok=True)

    # ── Data ────────────────────────────────────────────────────────────
    test_set = DarcyDataset(args.data_path, train=False,
                            reduced_resolution=args.reduced_res,
                            num_samples_max=args.num_samples)
    test_loader = DataLoader(test_set, batch_size=args.batch_size,
                             shuffle=False, pin_memory=True, num_workers=2)
    print(f"Test set: {len(test_set)} samples  ({test_set.H}×{test_set.W})\n")

    # ── Load models ─────────────────────────────────────────────────────
    print("Loading checkpoints ...")
    unet      = load_unet(args.ckpt_unet,      args.init_features, device)
    fno       = load_fno(args.ckpt_fno,        args.fno_modes, args.fno_width, device)
    pinn_unet = load_unet(args.ckpt_pinn_unet, args.init_features, device)
    print()

    model_registry = {
        "U-Net":      (unet,      "unet"),
        "FNO":        (fno,       "fno"),
        "PINN-U-Net": (pinn_unet, "unet"),
    }

    # ── Evaluate ─────────────────────────────────────────────────────────
    print("Evaluating models ...")
    results: dict[str, Optional[dict]] = {}
    for name, (model, mtype) in model_registry.items():
        if model is None:
            results[name] = None
            print(f"  {name:<14}  N/A (checkpoint not found)")
            continue
        print(f"  {name:<14}", end="", flush=True)
        met = evaluate_model(model, test_loader, device, mtype)
        results[name] = met
        print(f"  Rel-L2={met['rel_l2']:.4e}  RMSE={met['rmse']:.4e}  "
              f"MAE={met['mae']:.4e}  MaxErr={met['max_err']:.4e}  "
              f"PDE-Res={met['pde_residual_rmse']:.4e}  BC={met['boundary_err']:.4e}")

    # ── Terminal table ────────────────────────────────────────────────────
    col_w = 14
    header = f"\n{'Model':<{col_w}}" + "".join(
        f"{METRIC_LABELS[k]:>16}" for k in METRIC_KEYS
    )
    print(header)
    print("─" * len(header))
    for name, met in results.items():
        if met is None:
            print(f"{name:<{col_w}}" + "             N/A" * len(METRIC_KEYS))
        else:
            row = f"{name:<{col_w}}" + "".join(
                f"{met[k]:>16.4e}" for k in METRIC_KEYS
            )
            print(row)

    # ── Excel ──────────────────────────────────────────────────────────────
    print("\nExporting to Excel ...")
    export_excel(results, results_dir / "metrics_summary.xlsx")

    # ── Figure 2 — Training curves ─────────────────────────────────────────
    print("\nPlotting training curves ...")
    hist_paths = {
        "U-Net":      args.hist_unet,
        "FNO":        args.hist_fno,
        "PINN-U-Net": args.hist_pinn_unet,
    }
    save_training_curves(
        hist_paths,
        results_dir / "Figure_2_Training_Loss_and_Validation_Curves.pdf",
    )

    # ── Figure 3 — Bar chart ───────────────────────────────────────────────
    print("\nPlotting metrics bar chart ...")
    save_metrics_bar(results,
                     results_dir / "Figure_3_Model_Performance_Metrics_Comparison.pdf")

    # ── Figure 1 — 2-D field comparisons ──────────────────────────────────
    active_models = {k: v for k, v in model_registry.items() if v[0] is not None}
    if active_models:
        print(f"\nGenerating {args.num_vis} field comparison figure(s) ...")
        n_vis = min(args.num_vis, len(test_set))
        if args.vis_seed is not None:
            rng = np.random.default_rng(args.vis_seed)
            indices = sorted(rng.choice(len(test_set), size=n_vis, replace=False).tolist())
            print(f"  Random seed={args.vis_seed}, selected indices: {indices}")
        else:
            indices = list(range(n_vis))
        for idx in indices:
            sample_a, sample_u, sample_grid = test_set[idx]
            fname = (results_dir /
                     f"Figure_1_Darcy_Flow_Field_Comparison_Sample_{idx+1:02d}.pdf")
            save_field_comparison(
                active_models, sample_a, sample_u, sample_grid,
                device, idx, fname,
            )
    else:
        print("\n[skip] No valid models — field comparison figures not generated.")

    print(f"\nAll outputs saved to: {results_dir}/")


if __name__ == "__main__":
    main(get_args())
