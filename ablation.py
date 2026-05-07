"""
ablation.py — Ablation study for PINN-UNet design choices on 2D Darcy Flow.

Compares three PINN-UNet configurations:
  A) GradNorm only       (dynamic_lambda, no normalize_pde)
  B) NormPDE only        (normalize_pde,  no dynamic_lambda)
  C) NormPDE + GradNorm  (both enabled)

Outputs saved to ablation_results/:
  ablation_metrics.xlsx
  Ablation_A_Training_Curves.pdf / .png
  Ablation_B_Metrics_Bar.pdf / .png

Usage:
    python ablation.py
    python ablation.py --results_dir my_ablation
"""

from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Optional

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np
import torch
from torch.utils.data import DataLoader
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from matplotlib import font_manager as _fm

from models import UNet2d
from utils import DarcyDataset, compute_metrics

# ---------------------------------------------------------------------------
# Style
# ---------------------------------------------------------------------------

_has_times = any("Times New Roman" in f.name for f in _fm.fontManager.ttflist)
_SERIF = ["Times New Roman"] if _has_times else ["STIXGeneral", "DejaVu Serif"]

plt.rcParams.update({
    "font.family":       "serif",
    "font.serif":        _SERIF,
    "mathtext.fontset":  "stix",
    "font.size":         9,
    "axes.titlesize":    9,
    "axes.labelsize":    9,
    "xtick.labelsize":   8,
    "ytick.labelsize":   8,
    "legend.fontsize":   8,
    "savefig.dpi":       300,
    "axes.linewidth":    0.8,
    "figure.facecolor":  "white",
    "axes.facecolor":    "white",
})

METRIC_KEYS = ["rmse", "mae", "rel_l2", "max_err", "pde_residual_rmse", "boundary_err"]
METRIC_LABELS = {
    "rmse":              "RMSE",
    "mae":               "MAE",
    "rel_l2":            "Rel-$L_2$",
    "max_err":           "Max-Err ($L_\\infty$)",
    "pde_residual_rmse": "PDE-Res RMSE",
    "boundary_err":      "BC-RMSE",
}

# ---------------------------------------------------------------------------
# Ablation configurations
# ---------------------------------------------------------------------------

ABLATION_CONFIGS = [
    {
        "name":   "GradNorm only",
        "tag":    "pinn_gradnorm",
        "ckpt":   "checkpoints/low_data/pinn_gradnorm/best_model.pt",
        "hist":   "checkpoints/low_data/pinn_gradnorm/history.json",
        "color":  "#e08050",
        "marker": "o",
        "desc":   "dynamic lambda, raw residual",
    },
    {
        "name":   "NormPDE only",
        "tag":    "pinn_normpde",
        "ckpt":   "checkpoints/low_data/pinn_normpde/best_model.pt",
        "hist":   "checkpoints/low_data/pinn_normpde/history.json",
        "color":  "#5588cc",
        "marker": "s",
        "desc":   "fixed lambda=1.0, normalised residual",
    },
    {
        "name":   "NormPDE + GradNorm",
        "tag":    "pinn_norm_gradnorm",
        "ckpt":   "checkpoints/low_data/pinn_norm_gradnorm/best_model.pt",
        "hist":   "checkpoints/low_data/pinn_norm_gradnorm/history.json",
        "color":  "#55aa77",
        "marker": "^",
        "desc":   "dynamic lambda + normalised residual",
    },
]

# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def get_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="PINN-UNet ablation study")
    p.add_argument("--data_path",    type=str, default="dataset/2D_DarcyFlow_beta1.0_Train.hdf5")
    p.add_argument("--batch_size",   type=int, default=8)
    p.add_argument("--reduced_res",  type=int, default=1)
    p.add_argument("--init_features",type=int, default=32)
    p.add_argument("--results_dir",  type=str, default="ablation_results")
    return p.parse_args()

# ---------------------------------------------------------------------------
# Utilities
# ---------------------------------------------------------------------------

def load_model(ckpt_path: str, init_features: int, device) -> Optional[UNet2d]:
    p = Path(ckpt_path)
    if not p.exists():
        print(f"  [N/A] checkpoint not found: {ckpt_path}")
        return None
    model = UNet2d(in_channels=1, out_channels=1, init_features=init_features).to(device)
    sd = torch.load(ckpt_path, map_location=device, weights_only=True)
    sd = sd.get("model", sd)
    model.load_state_dict(sd)
    model.eval()
    return model


@torch.no_grad()
def evaluate_model(model: UNet2d, loader, device) -> dict[str, float]:
    accum   = {k: 0.0 for k in METRIC_KEYS + ["mse", "pde_residual"]}
    n_total = 0
    for a, u, _ in loader:
        a, u = a.to(device), u.to(device)
        u_pred = model(a.unsqueeze(1)).squeeze(1)
        m = compute_metrics(a, u_pred.float(), u)
        B = a.size(0)
        for k in accum:
            if k in m:
                accum[k] += m[k] * B
        n_total += B
    return {k: v / n_total for k, v in accum.items()}

# ---------------------------------------------------------------------------
# Excel export
# ---------------------------------------------------------------------------

def export_excel(results: dict, configs: list, save_path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ablation"

    hdr_font  = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    hdr_fill  = PatternFill("solid", fgColor="2F5496")
    row_fills = [PatternFill("solid", fgColor="D9E1F2"),
                 PatternFill("solid", fgColor="EBF0FA")]
    na_font   = Font(name="Calibri", italic=True, color="999999")
    center    = Alignment(horizontal="center", vertical="center")
    thin      = Side(style="thin", color="BFBFBF")
    border    = Border(left=thin, right=thin, top=thin, bottom=thin)

    col_headers = (["Configuration", "Description"] +
                   [METRIC_LABELS[k].replace("$", "").replace("\\", "") for k in METRIC_KEYS])
    for col, hdr in enumerate(col_headers, 1):
        cell = ws.cell(row=1, column=col, value=hdr)
        cell.font = hdr_font; cell.fill = hdr_fill
        cell.alignment = center; cell.border = border
        ws.column_dimensions[get_column_letter(col)].width = max(14, len(hdr) + 4)

    for row_i, cfg in enumerate(configs, 2):
        name = cfg["name"]
        met  = results.get(name)
        fill = row_fills[(row_i - 2) % 2]
        ws.cell(row=row_i, column=1, value=name).fill  = fill
        ws.cell(row=row_i, column=2, value=cfg["desc"]).fill = fill
        ws.cell(row=row_i, column=1).border = border
        ws.cell(row=row_i, column=2).border = border
        for col_j, k in enumerate(METRIC_KEYS, 3):
            cell = ws.cell(row=row_i, column=col_j)
            cell.fill = fill; cell.border = border; cell.alignment = center
            if met is None:
                cell.value = "N/A"; cell.font = na_font
            else:
                cell.value = round(met[k], 6)
                cell.number_format = "0.000000"

    wb.save(save_path)
    print(f"  Saved: {save_path.name}")

# ---------------------------------------------------------------------------
# Figure A — Training curves
# ---------------------------------------------------------------------------

def save_training_curves(configs: list, save_path: Path) -> None:
    fig, axes = plt.subplots(1, 2, figsize=(9, 3.8))
    fig.subplots_adjust(left=0.09, right=0.97, top=0.88, bottom=0.18, wspace=0.32)

    has_data = False
    for cfg in configs:
        p = Path(cfg["hist"])
        if not p.exists():
            continue
        try:
            hist = json.loads(p.read_text())
        except Exception:
            continue
        has_data = True
        kw = dict(color=cfg["color"], marker=cfg["marker"],
                  markersize=3, linewidth=1.2, label=cfg["name"])
        epochs = hist.get("epoch", [])
        if hist.get("train_mse"):
            axes[0].semilogy(epochs, hist["train_mse"], **kw)
        if hist.get("val_rel_l2"):
            axes[1].semilogy(epochs, hist["val_rel_l2"], **kw)

    if not has_data:
        for ax in axes:
            ax.text(0.5, 0.5, "No history found.", ha="center", va="center",
                    transform=ax.transAxes, color="gray")

    for ax, title, ylabel in zip(
        axes,
        ["(a) Training Loss", "(b) Validation Rel-$L_2$"],
        ["Training MSE Loss", "Relative $L_2$ Error"],
    ):
        ax.set_xlabel("Epoch"); ax.set_ylabel(ylabel); ax.set_title(title)
        ax.tick_params(which="both", direction="in")
        ax.grid(True, which="major", linestyle="--", linewidth=0.5, alpha=0.6)
        if has_data:
            ax.legend(framealpha=0.8)

    fig.suptitle("Ablation Study — PINN-UNet Training Progress",
                 fontsize=10, fontweight="bold")
    caption = (
        "Figure A. Training loss (MSE) and validation relative $L_2$ error for three PINN-UNet variants "
        "on 500 training samples. GradNorm: dynamic $\\lambda_{pde}$ via gradient-norm balancing. "
        "NormPDE: PDE residual divided by $(H-1)^2$ to remove grid-scale amplification."
    )
    fig.text(0.5, 0.01, caption, ha="center", va="bottom",
             fontsize=7, style="italic", transform=fig.transFigure)
    plt.savefig(save_path, dpi=300, bbox_inches="tight")
    plt.savefig(save_path.with_suffix(".png"), dpi=150, bbox_inches="tight")
    plt.close(fig)
    print(f"  Saved: {save_path.name}  +  {save_path.stem}.png")

# ---------------------------------------------------------------------------
# Figure B — Metrics bar chart
# ---------------------------------------------------------------------------

def save_metrics_bar(results: dict, configs: list, save_path: Path) -> None:
    valid = [c for c in configs if results.get(c["name"]) is not None]
    if not valid:
        print("  [skip] No valid results for bar chart.")
        return

    metrics_to_plot = ["rmse", "mae", "rel_l2", "max_err", "pde_residual_rmse", "boundary_err"]
    xlabels  = [METRIC_LABELS[k] for k in metrics_to_plot]
    n_models = len(valid)
    x        = np.arange(len(metrics_to_plot))
    width    = 0.22
    offsets  = np.linspace(-(n_models - 1) / 2, (n_models - 1) / 2, n_models) * width

    fig, ax = plt.subplots(figsize=(10, 4.5))
    fig.subplots_adjust(left=0.08, right=0.97, top=0.88, bottom=0.26)

    for i, cfg in enumerate(valid):
        met  = results[cfg["name"]]
        vals = [met.get(k, 0.0) for k in metrics_to_plot]
        bars = ax.bar(x + offsets[i], vals, width,
                      label=cfg["name"], color=cfg["color"],
                      alpha=0.85, edgecolor="white", linewidth=0.5)
        for bar, v in zip(bars, vals):
            ax.text(bar.get_x() + bar.get_width() / 2,
                    bar.get_height() * 1.10,
                    f"{v:.2e}",
                    ha="center", va="bottom", fontsize=6, rotation=90,
                    clip_on=False)

    ax.set_yscale("log")
    y_lo, y_hi = ax.get_ylim()
    ax.set_ylim(y_lo, y_hi * 8)
    ax.set_xticks(x)
    ax.set_xticklabels(xlabels, rotation=30, ha="right", fontsize=8)
    ax.set_ylabel("Metric Value (log scale)")
    ax.set_title("Ablation Study — PINN-UNet Configuration Comparison",
                 fontsize=10, fontweight="bold")
    ax.legend(framealpha=0.85, fontsize=8)
    ax.tick_params(which="both", direction="in")
    ax.grid(True, axis="y", which="major", linestyle="--", linewidth=0.5, alpha=0.6)

    caption = (
        "Figure B. Ablation comparison of three PINN-UNet variants on the Darcy flow test set "
        "($N=1000$, 500 training samples). "
        "GradNorm only: dynamic $\\lambda$ without residual normalisation. "
        "NormPDE only: fixed $\\lambda=1.0$ with $(H-1)^{-2}$ normalisation. "
        "NormPDE + GradNorm: both techniques combined. Lower is better."
    )
    fig.text(0.5, 0.01, caption, ha="center", va="bottom",
             fontsize=7, style="italic", transform=fig.transFigure)

    plt.savefig(save_path, dpi=300, bbox_inches="tight")
    plt.savefig(save_path.with_suffix(".png"), dpi=150, bbox_inches="tight")
    plt.close(fig)
    print(f"  Saved: {save_path.name}  +  {save_path.stem}.png")

# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main(args: argparse.Namespace) -> None:
    device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
    print(f"\nDevice: {device}")

    results_dir = Path(args.results_dir)
    results_dir.mkdir(parents=True, exist_ok=True)

    test_set = DarcyDataset(args.data_path, train=False,
                            reduced_resolution=args.reduced_res)
    test_loader = DataLoader(test_set, batch_size=args.batch_size,
                             shuffle=False, pin_memory=True, num_workers=2)
    print(f"Test set: {len(test_set)} samples  ({test_set.H}x{test_set.W})\n")

    results: dict[str, Optional[dict]] = {}
    print("Loading and evaluating configurations ...")
    for cfg in ABLATION_CONFIGS:
        print(f"  {cfg['name']:<25}", end="", flush=True)
        model = load_model(cfg["ckpt"], args.init_features, device)
        if model is None:
            results[cfg["name"]] = None
            print("  N/A")
            continue
        met = evaluate_model(model, test_loader, device)
        results[cfg["name"]] = met
        print(f"  Rel-L2={met['rel_l2']:.4e}  RMSE={met['rmse']:.4e}"
              f"  PDE-Res={met['pde_residual_rmse']:.4e}  BC={met['boundary_err']:.4e}")

    col_w  = 25
    header = f"\n{'Configuration':<{col_w}}" + "".join(
        f"{METRIC_LABELS[k]:>16}" for k in METRIC_KEYS)
    print(header)
    print("─" * len(header))
    for cfg in ABLATION_CONFIGS:
        name = cfg["name"]
        met  = results.get(name)
        if met is None:
            print(f"{name:<{col_w}}" + "             N/A" * len(METRIC_KEYS))
        else:
            print(f"{name:<{col_w}}" + "".join(f"{met[k]:>16.4e}" for k in METRIC_KEYS))

    print("\nExporting to Excel ...")
    export_excel(results, ABLATION_CONFIGS, results_dir / "ablation_metrics.xlsx")

    print("\nPlotting training curves ...")
    save_training_curves(ABLATION_CONFIGS, results_dir / "Ablation_A_Training_Curves.pdf")

    print("\nPlotting metrics bar chart ...")
    save_metrics_bar(results, ABLATION_CONFIGS, results_dir / "Ablation_B_Metrics_Bar.pdf")

    print(f"\nAll ablation outputs saved to: {results_dir}/\n")


if __name__ == "__main__":
    main(get_args())
